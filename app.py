# app.py
import os, re, datetime, csv, sqlite3, json, math
from copy import copy
from functools import wraps

from flask import (
    Flask, request, render_template, jsonify, send_file,
    redirect, url_for, session, abort
)

import pdfplumber
import pandas as pd
import openpyxl
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

# =========================
# PATHS ABSOLUTOS (evita problemas en server)
# =========================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))

def p(*parts):  # helper path join
    return os.path.join(BASE_DIR, *parts)

# =========================
# CONFIG FACTURACIÓN
# =========================
OPERATIVAS = {
    "PATIO_ECI": {
        "label": "PATIO ECI (Valdemoro)",
        "template_xlsx": p("plantilla patio.xlsx"),
        "master_xlsx": p("maestro_matriculas.xlsx"),          # Conductor -> Matricula/Ruta
        "vehiculos_xlsx": p("maestro_vehiculos.xlsx"),        # Matricula -> Proveedor
        "default_ruta": "V429",
        "cobro_ruta": {"V429": 32, "V429.1": 48, "V429.2": 29},
        "rutas": ["V429", "V429.1", "V429.2"],
        "cliente_codigo": 2,              # Col A
        "remolque_ref": "M111111",        # Col AI
        "proyecto": "V429",               # Col I (por ahora fijo)
    }
}
DEFAULT_OPERATIVA = "PATIO_ECI"
OUTPUT_XLSX = p("salida_meribia.xlsx")

KPI_FILE = p("kpis_facturacion.csv")
PROVEEDORES_FILE = p("proveedores_master.csv")

# =========================
# NUEVO: LINKS FLOTA
# =========================
FLOTA_SHEET_URL = "https://docs.google.com/spreadsheets/d/1mdK6gKjBpDF7vFD1R54bu9GCKuQJuxGJK7OMKsKGoHg/edit?gid=0#gid=0"
FLOTA_LISTIN_URL = "https://script.google.com/macros/s/AKfycbzoSiZo757K3CuCIz0aEmWJX2idaIWUqwVl5rA6MZsT9npyf5zZzb_6UZ7lhun3a_Krcg/exec?viewer=1"

from routes_module import rutas_bp
# =========================
# MÓDULOS DISPONIBLES (permisos)
# =========================
MODULES = [
    {"id": "realtime", "label": "Control Panel (Tiempo real)"},
    {"id": "facturacion_patio", "label": "Facturación Patio ECI → Meribia"},
    {"id": "flota", "label": "Inventario Flota (Listín + Maestro)"},
    {"id": "personal", "label": "Personal (Conductores / Operativos)"},
    # =========================
    # ✅ NUEVOS (NO TOCAN módulos existentes)
    # =========================
    {"id": "rutas", "label": "Rutas (ECI + otros)"},
    {"id": "planificacion", "label": "Planificación (borradores / plantillas)"},
    {"id": "clientes", "label": "Clientes (fichas + contactos)"},
    {"id": "proveedores_fichas", "label": "Proveedores (fichas + contactos)"},
    {"id": "vehiculos", "label": "Vehículos (docs, ITV, tacógrafo, ATP...)"},
]

# =========================
# COLUMNAS (1-index) MERIBIA (plantilla patio.xlsx)
# =========================
COL_A_CLIENTE = 1
COL_I_PROYECTO = 9
COL_R_FECHA_DESCARGA = 18
COL_AE_HORAS_REALES = 31
COL_AF_PRECIO_CLIENTE = 32
COL_AH_MATRICULA = 34
COL_AI_REMOLQUE = 35
COL_AL_HORAS_COSTE = 38
COL_AM_PRECIO_UNI = 39
COL_AN_IMPORTE = 40

# =========================
# PROVEEDORES POR DEFECTO (facturación patio)
# =========================
PROVEEDORES_DEFAULT = {
    "MARTIN SIMANCAS": {"tipo": "hora", "pago_h": 25.0, "pago_f": 30.0},
    "JUAN CALVO":      {"tipo": "hora", "pago_h": 25.0, "pago_f": 30.0},
    "ARANDA":          {"tipo": "hora", "pago_h": 25.0, "pago_f": 30.0},
    "CANELO":          {"tipo": "hora", "pago_h": 25.0, "pago_f": 30.0},
    "ANGEL MUNOZ":     {"tipo": "hora", "pago_h": 22.5, "pago_f": 28.0},
    "TRANSMAU":        {"tipo": "hora", "pago_h": 25.0, "pago_f": 30.0},
    "PIBEJO":          {"tipo": "hora", "pago_h": 0.0,  "pago_f": 0.0},
    "CAMPOY":          {"tipo": "hora", "pago_h": 25.0, "pago_f": 30.0},
    "ALBERTO RAMAL":   {"tipo": "hora", "pago_h": 25.0, "pago_f": 30.0},
    "RUBEN CUESTA":    {"tipo": "dia",  "pago_dia": 260.0, "pago_dia_f": 275.0},
}

# =========================
# FLASK APP
# =========================
app = Flask(__name__)
app.register_blueprint(rutas_bp)
app.config["UPLOAD_FOLDER"] = p("uploads")
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

app.config["MAX_CONTENT_LENGTH"] = 25 * 1024 * 1024
app.secret_key = os.environ.get("FLASK_SECRET_KEY", "CAMBIA-ESTO-POR-UNA-FRASE-LARGA-123456")

# =========================
# USUARIOS (SQLite)
# =========================
DB_USERS = p("users.db")

def db():
    conn = sqlite3.connect(DB_USERS)
    conn.row_factory = sqlite3.Row
    return conn

def init_users_db():
    with db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users(
                username TEXT PRIMARY KEY,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL DEFAULT 'user',
                modules TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL
            )
        """)
        try:
            conn.execute("ALTER TABLE users ADD COLUMN modules TEXT NOT NULL DEFAULT ''")
        except Exception:
            pass
        conn.commit()

def ensure_default_admin():
    username = "admin"
    password = "tazjcab55"
    with db() as conn:
        r = conn.execute("SELECT username FROM users WHERE username=?", (username,)).fetchone()
        if not r:
            conn.execute(
                "INSERT INTO users(username,password_hash,role,modules,created_at) VALUES(?,?,?,?,?)",
                (
                    username,
                    generate_password_hash(password),
                    "admin",
                    "",
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                )
            )
            conn.commit()

def get_user(username: str):
    with db() as conn:
        return conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()

def csv_to_modules(s: str):
    s = (s or "").strip()
    if not s:
        return []
    return [x.strip() for x in s.split(",") if x.strip()]

def modules_to_csv(mods: list):
    mods = [m.strip() for m in (mods or []) if m and m.strip()]
    valid = {m["id"] for m in MODULES}
    mods = [m for m in mods if m in valid]
    out = []
    seen = set()
    for m in mods:
        if m not in seen:
            out.append(m)
            seen.add(m)
    return ",".join(out)

def list_users():
    with db() as conn:
        rows = conn.execute(
            "SELECT username, role, modules, created_at FROM users ORDER BY role DESC, username ASC"
        ).fetchall()
        out = []
        for r in rows:
            d = dict(r)
            d["modules_list"] = csv_to_modules(d.get("modules", ""))
            out.append(d)
        return out

def create_user(username: str, password: str, role: str, modules: list):
    username = (username or "").strip()
    if not username:
        raise ValueError("Usuario vacío")
    if len(password or "") < 6:
        raise ValueError("La contraseña debe tener al menos 6 caracteres")

    role = role if role in ("admin", "user", "driver") else "user"
    modules_csv = "" if role in ("admin", "driver") else modules_to_csv(modules)

    with db() as conn:
        if conn.execute("SELECT username FROM users WHERE username=?", (username,)).fetchone():
            raise ValueError("Ese usuario ya existe")
        conn.execute(
            "INSERT INTO users(username,password_hash,role,modules,created_at) VALUES(?,?,?,?,?)",
            (
                username,
                generate_password_hash(password),
                role,
                modules_csv,
                datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            )
        )
        conn.commit()

def set_password(username: str, password: str):
    username = (username or "").strip()
    if not username:
        raise ValueError("Usuario vacío")
    if len(password or "") < 6:
        raise ValueError("La contraseña debe tener al menos 6 caracteres")

    with db() as conn:
        if not conn.execute("SELECT username FROM users WHERE username=?", (username,)).fetchone():
            raise ValueError("Ese usuario no existe")
        conn.execute(
            "UPDATE users SET password_hash=? WHERE username=?",
            (generate_password_hash(password), username)
        )
        conn.commit()

def set_modules(username: str, modules: list):
    username = (username or "").strip()
    if not username:
        raise ValueError("Usuario vacío")
    if username == "admin":
        return

    user = get_user(username)
    if user and user["role"] == "driver":
        return

    modules_csv = modules_to_csv(modules)
    with db() as conn:
        if not conn.execute("SELECT username FROM users WHERE username=?", (username,)).fetchone():
            raise ValueError("Ese usuario no existe")
        conn.execute("UPDATE users SET modules=? WHERE username=?", (modules_csv, username))
        conn.commit()

def delete_user(username: str):
    username = (username or "").strip()
    if username == "admin":
        raise ValueError("No se puede borrar el usuario admin")
    with db() as conn:
        conn.execute("DELETE FROM users WHERE username=?", (username,))
        conn.commit()

init_users_db()
ensure_default_admin()

# =========================
# AUTH DECORATORS
# =========================
def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        return fn(*args, **kwargs)
    return wrapper

def admin_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get("logged_in"):
            return redirect(url_for("login"))
        if session.get("role") != "admin":
            abort(403)
        return fn(*args, **kwargs)
    return wrapper

def module_required(module_id: str):
    def deco(fn):
        @wraps(fn)
        def wrapper(*args, **kwargs):
            if not session.get("logged_in"):
                return redirect(url_for("login"))
            if session.get("role") == "admin":
                return fn(*args, **kwargs)

            allowed = session.get("modules_list", []) or []
            if module_id not in allowed:
                abort(403)
            return fn(*args, **kwargs)
        return wrapper
    return deco

# =========================
# ERROR HANDLER 403
# =========================
@app.errorhandler(403)
def forbidden(_):
    return render_template("forbidden.html", module=request.path), 403

# =========================
# LOGIN ROUTES
# =========================
@app.route("/login", methods=["GET", "POST"])
def login():
    init_users_db()
    ensure_default_admin()

    if request.method == "POST":
        u = (request.form.get("username", "") or "").strip()
        p_ = (request.form.get("password", "") or "").strip()

        user = get_user(u)
        if user and check_password_hash(user["password_hash"], p_):
            session.clear()
            session["logged_in"] = True
            session["username"] = user["username"]
            session["role"] = user["role"]

            mods = csv_to_modules(user["modules"])
            session["modules_csv"] = user["modules"] or ""
            session["modules_list"] = mods

            return redirect(url_for("dashboard"))

        return render_template("login.html", error="Usuario o contraseña incorrectos.")

    return render_template("login.html", error=None)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

# =========================
# HELPERS FACTURACIÓN (existente)
# =========================
def parse_spanish_number(s: str) -> float:
    s = str(s).strip().replace(".", "").replace(",", ".")
    return float(s)

def normalize_line(line: str) -> str:
    return re.sub(r"\s+", " ", line).strip()

def cleanup_numbers(line: str) -> str:
    s = re.sub(r"(\d)\s*([,\.])\s*(\d)", r"\1\2\3", line)
    return re.sub(r"\s+", " ", s).strip()

def key_name(raw: str) -> str:
    s = str(raw).upper()
    s = re.sub(r"[^A-ZÁÉÍÓÚÜÑ0-9 ]", "", s)
    s = s.translate(str.maketrans("ÁÉÍÓÚÜÑ", "AEIOUUN"))
    return re.sub(r"\s+", " ", s).strip()

def key_plate(raw: str) -> str:
    s = str(raw).upper().strip()
    s = re.sub(r"[^A-Z0-9]", "", s)
    return s

def parse_pdf_line_flex(line: str):
    m = re.search(r"(\d+[.,]\d+|\d+)\s+(\d+[.,]\d+|\d+)\s+(\d+[.,]\d+|\d+)\s*$", line)
    if m:
        horas_reales = parse_spanish_number(m.group(3))
        core = line[:m.start()].strip()
    else:
        m2 = re.search(r"(\d+[.,]\d+|\d+)\s+(\d+[.,]\d+|\d+)\s*$", line)
        if not m2:
            return None
        horas_reales = parse_spanish_number(m2.group(2))
        core = line[:m2.start()].strip()

    if "Diaria" not in core:
        return None

    _, post = core.split("Diaria", 1)
    parts = post.strip().split()
    rest = " ".join(parts[1:]).strip() if parts else ""
    return {"rest": rest, "horas_reales": float(horas_reales)}

def split_conductor_transportista(rest: str):
    up = rest.upper()
    m = re.search(r"\bTRANS(?:PORTES)?\b|\bTRANSPORTES\b|\bTRANSPORTE\b", up)
    if m:
        return rest[:m.start()].strip(), rest[m.start():].strip()
    return rest.strip(), rest.strip()

# =========================
# PROVEEDORES (CSV) (existente)
# =========================
def ensure_proveedores_file():
    if os.path.exists(PROVEEDORES_FILE):
        return
    with open(PROVEEDORES_FILE, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["nombre", "tipo", "pago_h", "pago_f", "pago_dia", "pago_dia_f"])

def load_proveedores() -> dict:
    prov = dict(PROVEEDORES_DEFAULT)
    if not os.path.exists(PROVEEDORES_FILE):
        return prov
    try:
        df = pd.read_csv(PROVEEDORES_FILE, sep=";")
        for _, r in df.iterrows():
            nombre = str(r.get("nombre", "")).strip()
            if not nombre:
                continue
            nombre = key_name(nombre)
            tipo = str(r.get("tipo", "hora")).strip().lower()
            item = {"tipo": tipo}
            if tipo == "dia":
                item["pago_dia"] = float(r.get("pago_dia", 0) or 0)
                item["pago_dia_f"] = float(r.get("pago_dia_f", 0) or 0)
            else:
                item["pago_h"] = float(r.get("pago_h", 0) or 0)
                item["pago_f"] = float(r.get("pago_f", 0) or 0)
            prov[nombre] = item
    except Exception:
        pass
    return prov

def upsert_proveedor(nombre: str, data: dict):
    ensure_proveedores_file()
    nombre_norm = key_name(nombre)

    tipo = str(data.get("tipo", "hora")).strip().lower()
    if tipo not in ("hora", "dia"):
        tipo = "hora"

    if os.path.exists(PROVEEDORES_FILE):
        df = pd.read_csv(PROVEEDORES_FILE, sep=";")
    else:
        df = pd.DataFrame(columns=["nombre","tipo","pago_h","pago_f","pago_dia","pago_dia_f"])

    if "nombre" not in df.columns:
        df["nombre"] = ""

    df["nombre_norm"] = df["nombre"].astype(str).apply(key_name)
    mask = df["nombre_norm"] == nombre_norm

    if tipo == "dia":
        row = {
            "nombre": nombre_norm,
            "tipo": "dia",
            "pago_h": "",
            "pago_f": "",
            "pago_dia": float(data.get("pago_dia", 0) or 0),
            "pago_dia_f": float(data.get("pago_dia_f", 0) or 0),
        }
    else:
        row = {
            "nombre": nombre_norm,
            "tipo": "hora",
            "pago_h": float(data.get("pago_h", 0) or 0),
            "pago_f": float(data.get("pago_f", 0) or 0),
            "pago_dia": "",
            "pago_dia_f": "",
        }

    if mask.any():
        idx = df.index[mask][0]
        for k, v in row.items():
            df.at[idx, k] = v
    else:
        df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)

    df = df.drop(columns=["nombre_norm"], errors="ignore")
    df.to_csv(PROVEEDORES_FILE, sep=";", index=False, encoding="utf-8")

# =========================
# MATRÍCULA -> PROVEEDOR (maestro_vehiculos.xlsx) (existente)
# =========================
def ensure_vehiculos_master(path: str):
    if os.path.exists(path):
        return
    df = pd.DataFrame(columns=["Matricula", "Proveedor"])
    df.to_excel(path, index=False)

def load_vehiculo_map(path: str) -> dict:
    if not os.path.exists(path):
        return {}
    df = pd.read_excel(path)
    if "Matricula" not in df.columns or "Proveedor" not in df.columns:
        return {}
    mp = {}
    for _, r in df.iterrows():
        m = key_plate(r.get("Matricula", ""))
        pr = key_name(r.get("Proveedor", ""))
        if m:
            mp[m] = pr
    return mp

def save_vehiculo_map(path: str, rows: list):
    ensure_vehiculos_master(path)
    df = pd.read_excel(path)

    if "Matricula" not in df.columns:
        df["Matricula"] = ""
    if "Proveedor" not in df.columns:
        df["Proveedor"] = ""

    df["MatKey"] = df["Matricula"].astype(str).apply(key_plate)
    idx = {k: i for i, k in enumerate(df["MatKey"].tolist())}

    for row in rows:
        m = key_plate(row.get("Matricula", ""))
        pr = key_name(row.get("Proveedor", ""))
        if not m or not pr:
            continue
        if m in idx:
            df.at[idx[m], "Proveedor"] = pr
        else:
            df = pd.concat([df, pd.DataFrame([{"Matricula": m, "Proveedor": pr, "MatKey": m}])], ignore_index=True)
            idx[m] = len(df) - 1

    df = df.drop(columns=["MatKey"], errors="ignore")
    df.to_excel(path, index=False)

# =========================
# PARSEO PDF + AGRUPACIÓN (facturación patio) (existente)
# =========================
def consignatario_guess_from_transportista(tr: str) -> str:
    up = str(tr).upper()
    if "PIBEJO" in up: return "PIBEJO"
    if "CAMPOY" in up: return "CAMPOY"
    if "SIMANCAS" in up: return "MARTIN SIMANCAS"
    if "ARANDA" in up: return "ARANDA"
    if "CALVO" in up: return "JUAN CALVO"
    if "TRANSMAU" in up or "TRANS MAU" in up: return "TRANSMAU"
    if "ANGEL" in up and ("MUNOZ" in up or "MUÑOZ" in up): return "ANGEL MUNOZ"
    if "RUBEN" in up or "CUESTA" in up: return "RUBEN CUESTA"
    return ""

def parse_and_group(pdf_path: str):
    lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for i, page in enumerate(pdf.pages):
            try:
                txt = page.extract_text(x_tolerance=2, y_tolerance=2) or ""
                if not txt.strip():
                    txt = page.extract_text_simple() or ""
            except Exception as e:
                print(f"[WARN] PDF page {i+1} parse failed: {e}")
                txt = ""

            for ln in (txt.splitlines() if txt else []):
                ln = ln.strip()
                if not ln:
                    continue
                if ln.startswith("FechaInicioJornada"):
                    continue
                lines.append(ln)

    rows = []
    for ln in lines:
        ln2 = cleanup_numbers(normalize_line(ln))
        d = parse_pdf_line_flex(ln2)
        if not d:
            continue

        conductor, transportista = split_conductor_transportista(d["rest"])
        consignatario = consignatario_guess_from_transportista(transportista) or transportista

        rows.append({
            "Conductor": conductor,
            "Transportista": transportista,
            "Proveedor": key_name(consignatario) if consignatario else "",
            "HorasReales": float(d["horas_reales"]),
        })

    if not rows:
        return []

    df = pd.DataFrame(rows)
    df["ConductorKey"] = df["Conductor"].apply(key_name)

    grouped = (df.groupby("ConductorKey")
        .agg(
            Conductor=("Conductor", "first"),
            Transportista=("Transportista", lambda s: s.value_counts().index[0]),
            Proveedor=("Proveedor", lambda s: s.value_counts().index[0] if len(s.value_counts()) else ""),
            HorasReales=("HorasReales", "sum"),
            Registros=("HorasReales", "count"),
        )
        .reset_index()
        .sort_values("Conductor")
    )
    return grouped.to_dict(orient="records")

# =========================
# MAESTRO Conductor -> Matricula/Ruta (existente)
# =========================
def ensure_master_exists(master_path: str):
    if os.path.exists(master_path):
        return
    df = pd.DataFrame(columns=["Conductor", "Matricula", "Ruta"])
    df.to_excel(master_path, index=False)

def load_master_map(master_path: str):
    if not os.path.exists(master_path):
        return {}
    df = pd.read_excel(master_path)
    cols = {c.lower().strip(): c for c in df.columns}
    if "conductor" not in cols or "matricula" not in cols:
        return {}
    c_conductor = cols["conductor"]
    c_matricula = cols["matricula"]
    c_ruta = cols.get("ruta")

    mp = {}
    for _, row in df.iterrows():
        name = str(row[c_conductor]).strip()
        if not name or name.lower() == "nan":
            continue
        k = key_name(name)
        mp[k] = {
            "Matricula": str(row[c_matricula]).strip() if str(row[c_matricula]).lower() != "nan" else "",
            "Ruta": (str(row[c_ruta]).strip() if c_ruta and str(row[c_ruta]).lower() != "nan" else "")
        }
    return mp

def apply_master(rows, master_map, default_ruta):
    for r in rows:
        k = r.get("ConductorKey") or key_name(r.get("Conductor", ""))
        r["MatriculaFromMaster"] = False
        if k in master_map:
            if not r.get("Matricula") and master_map[k].get("Matricula"):
                r["Matricula"] = master_map[k].get("Matricula", "")
                r["MatriculaFromMaster"] = True
            if not r.get("Ruta") and master_map[k].get("Ruta"):
                r["Ruta"] = master_map[k].get("Ruta", default_ruta)
        if not r.get("Ruta"):
            r["Ruta"] = default_ruta
    return rows

def save_master_from_rows(master_path: str, rows: list):
    ensure_master_exists(master_path)
    df_old = pd.read_excel(master_path)

    if "Ruta" not in df_old.columns:
        df_old["Ruta"] = ""

    df_old["Key"] = df_old["Conductor"].astype(str).apply(key_name)
    idx = {k: i for i, k in enumerate(df_old["Key"].tolist())}

    for r in rows:
        name = str(r.get("Conductor", "")).strip()
        if not name:
            continue
        k = key_name(name)
        mat = str(r.get("Matricula", "")).strip()
        ruta = str(r.get("Ruta", "")).strip()

        if k in idx:
            i = idx[k]
            if mat:
                df_old.at[i, "Matricula"] = mat
            if ruta:
                df_old.at[i, "Ruta"] = ruta
        else:
            df_old = pd.concat([df_old, pd.DataFrame([{
                "Conductor": name,
                "Matricula": mat,
                "Ruta": ruta,
                "Key": k
            }])], ignore_index=True)
            idx[k] = len(df_old) - 1

    df_old = df_old.drop(columns=["Key"], errors="ignore")
    df_old.to_excel(master_path, index=False)

# =========================
# COSTES (existente)
# =========================
def is_propio(proveedor: str) -> bool:
    return key_name(proveedor) == "PIBEJO"

def compute_cost_row(row: dict, es_festivo: bool, prov_map: dict):
    if bool(row.get("OverrideCoste")) and row.get("CosteManual") is not None:
        manual = float(row.get("CosteManual") or 0.0)
        return {"qty": 0.0, "unit": 0.0, "importe": manual, "tipo": "manual"}

    horas = float(row.get("HorasReales", 0) or 0)
    proveedor = key_name(row.get("Proveedor", "") or "")

    if is_propio(proveedor):
        return {"qty": 0.0, "unit": 0.0, "importe": 0.0, "tipo": "propio"}

    info = prov_map.get(proveedor)
    if not info:
        return {"qty": 0.0, "unit": 0.0, "importe": 0.0, "tipo": "sin_tarifa"}

    tipo = info.get("tipo", "hora")
    if tipo == "dia":
        unit = float(info.get("pago_dia_f" if es_festivo else "pago_dia", 0) or 0)
        return {"qty": 1.0, "unit": unit, "importe": unit, "tipo": "dia"}

    unit = float(info.get("pago_f" if es_festivo else "pago_h", 0) or 0)
    return {"qty": horas, "unit": unit, "importe": horas * unit, "tipo": "hora"}

# =========================
# EXPORT MERIBIA (existente)
# =========================
def generate_meribia_xlsx(
    rows, date_iso: str, template_xlsx: str, es_festivo: bool, prov_map: dict,
    cliente_codigo: int, remolque_ref: str, proyecto: str, cobro_ruta: dict
):
    fecha = datetime.date.fromisoformat(date_iso)

    wb = openpyxl.load_workbook(template_xlsx)
    if "PLANTILLA" not in wb.sheetnames:
        raise ValueError("No existe la hoja 'PLANTILLA' en la plantilla.")
    ws = wb["PLANTILLA"]
    max_col = ws.max_column

    for rr in range(2, ws.max_row + 1):
        for cc in range(1, max_col + 1):
            ws.cell(rr, cc).value = None

    needed = len(rows) + 1
    if ws.max_row < needed:
        ws.insert_rows(ws.max_row + 1, amount=needed - ws.max_row)

    def copy_row_style(src_r, tgt_r):
        for c in range(1, max_col + 1):
            src = ws.cell(src_r, c)
            tgt = ws.cell(tgt_r, c)
            tgt._style = copy(src._style)
            tgt.number_format = src.number_format
            tgt.font = copy(src.font)
            tgt.border = copy(src.border)
            tgt.fill = copy(src.fill)
            tgt.alignment = copy(src.alignment)
            tgt.protection = copy(src.protection)

    if ws.max_row >= 2:
        for i in range(len(rows)):
            copy_row_style(2, 2 + i)

    total_coste = 0.0

    for i, row in enumerate(rows):
        r = 2 + i
        ruta = str(row.get("Ruta", "V429")).strip()
        horas = float(row.get("HorasReales", 0) or 0)
        proveedor = key_name(row.get("Proveedor", "") or "")

        cost = compute_cost_row(row, es_festivo, prov_map)

        ws.cell(r, COL_A_CLIENTE).value = int(cliente_codigo)
        ws.cell(r, COL_I_PROYECTO).value = str(proyecto)
        ws.cell(r, COL_R_FECHA_DESCARGA).value = fecha

        ws.cell(r, 4).value = fecha
        ws.cell(r, 10).value = ruta
        ws.cell(r, 2).value = proveedor

        ws.cell(r, COL_AE_HORAS_REALES).value = horas
        ws.cell(r, COL_AF_PRECIO_CLIENTE).value = float(cobro_ruta.get(ruta, 0) or 0)
        ws.cell(r, COL_AH_MATRICULA).value = row.get("Matricula", "")
        ws.cell(r, COL_AI_REMOLQUE).value = str(remolque_ref)

        ws.cell(r, COL_AL_HORAS_COSTE).value = float(cost["qty"])
        ws.cell(r, COL_AM_PRECIO_UNI).value = float(cost["unit"])
        ws.cell(r, COL_AN_IMPORTE).value = None

        ws.cell(r, 33).value = f"Chofer: {row.get('Conductor','')}"
        total_coste += float(cost["importe"])

    total_horas = sum(float(x.get("HorasReales", 0) or 0) for x in rows)

    wb.save(OUTPUT_XLSX)
    return OUTPUT_XLSX, float(total_horas), float(total_coste)

# =========================
# KPI (existente)
# =========================
def append_kpi(date_iso: str, operativa: str, rows: list, cobro_ruta: dict,
               rutas_validas: list, es_festivo: bool, total_horas: float, total_coste: float):
    total_cobro = 0.0
    for r in rows:
        ruta = str(r.get("Ruta", "V429")).strip()
        horas = float(r.get("HorasReales", 0) or 0)
        tarifa = float(cobro_ruta.get(ruta, 0) or 0)
        total_cobro += horas * tarifa
    total_cobro = round(total_cobro, 2)

    manual_count = sum(1 for r in rows if bool(r.get("OverrideCoste")))
    conductores = len(rows)

    file_exists = os.path.exists(KPI_FILE)
    with open(KPI_FILE, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f, delimiter=";")
        if not file_exists:
            w.writerow(["fecha","operativa","festivo","conductores","horas_total","cobro_total","coste_total","manual_costes"])
        w.writerow([date_iso, operativa, int(es_festivo), conductores, round(total_horas,2), total_cobro, round(total_coste,2), manual_count])

# =========================
# ✅ PERSONAL IMPORT (existente)
# =========================
def read_people_file(file_storage):
    filename = (file_storage.filename or "").lower()
    if filename.endswith(".xlsx") or filename.endswith(".xls"):
        df = pd.read_excel(file_storage)
    else:
        content = file_storage.read()
        file_storage.seek(0)
        try:
            df = pd.read_csv(file_storage, sep=None, engine="python")
        except Exception:
            file_storage.seek(0)
            df = pd.read_csv(file_storage, sep=";")

    df.columns = [str(c).strip().lower() for c in df.columns]
    return df

def import_drivers_from_df(df: pd.DataFrame):
    required = ["username", "pin"]
    for c in required:
        if c not in df.columns:
            raise ValueError(f"Falta columna obligatoria: {c}")

    created = 0
    skipped = 0
    errors = []

    for idx, row in df.iterrows():
        try:
            username = str(row.get("username", "")).strip()
            pin = str(row.get("pin", "")).strip()
            if not username or not pin:
                skipped += 1
                continue
            if len(pin) < 6:
                raise ValueError("PIN mínimo 6 dígitos/caracteres")

            if get_user(username):
                skipped += 1
                continue

            create_user(username=username, password=pin, role="driver", modules=[])

            created += 1
        except Exception as e:
            errors.append(f"Fila {idx+2}: {e}")

    return created, skipped, errors

# =========================
# PORTAL PAGES (existente)
# =========================
@app.route("/")
@login_required
def dashboard():
    is_admin = (session.get("role") == "admin")
    if is_admin:
        modules = MODULES[:]
    else:
        allowed = set(session.get("modules_list", []) or [])
        modules = [m for m in MODULES if m["id"] in allowed]
    return render_template("dashboard.html", modules=modules, is_admin=is_admin)

@app.route("/panel/realtime")
@login_required
@module_required("realtime")
def panel_realtime():
    return render_template("panel_realtime.html")

@app.route("/facturacion/patio")
@login_required
@module_required("facturacion_patio")
def facturacion_patio():
    ops = [{"id": k, "label": v["label"]} for k, v in OPERATIVAS.items()]
    return render_template("patio.html", operativas=ops, default_operativa=DEFAULT_OPERATIVA)

@app.route("/flota")
@login_required
@module_required("flota")
def flota_home():
    return render_template("flota.html", sheet_url=FLOTA_SHEET_URL, listin_url=FLOTA_LISTIN_URL)

@app.route("/flota/sheet")
@login_required
@module_required("flota")
def flota_sheet():
    return redirect(FLOTA_SHEET_URL)

@app.route("/flota/listin")
@login_required
@module_required("flota")
def flota_listin():
    return redirect(FLOTA_LISTIN_URL)

@app.route("/personal")
@login_required
@module_required("personal")
def personal_home():
    is_admin = (session.get("role") == "admin")
    return render_template("personal.html", is_admin=is_admin, ok=None, error=None)

@app.route("/personal/import", methods=["POST"])
@admin_required
def personal_import():
    try:
        f = request.files.get("file")
        if not f:
            raise ValueError("No se recibió archivo")

        df = read_people_file(f)
        created, skipped, errors = import_drivers_from_df(df)

        ok = f"Importación finalizada: creados {created}, omitidos {skipped}."
        if errors:
            ok += f" Con {len(errors)} errores (revísalos abajo)."

        return render_template("personal.html", is_admin=True, ok=ok, error="\n".join(errors) if errors else None)
    except Exception as e:
        return render_template("personal.html", is_admin=True, ok=None, error=str(e))

# =========================
# ADMIN USERS PANEL (existente)
# =========================
@app.route("/admin/users")
@admin_required
def admin_users():
    return render_template("users.html", users=list_users(), modules=MODULES, ok=None, error=None)

@app.route("/admin/users/create", methods=["POST"])
@admin_required
def admin_users_create():
    try:
        modules = request.form.getlist("modules")
        create_user(
            request.form.get("username",""),
            request.form.get("password",""),
            request.form.get("role","user"),
            modules
        )
        return render_template("users.html", users=list_users(), modules=MODULES, ok="Usuario creado.", error=None)
    except Exception as e:
        return render_template("users.html", users=list_users(), modules=MODULES, ok=None, error=str(e))

@app.route("/admin/users/modules", methods=["POST"])
@admin_required
def admin_users_modules():
    try:
        username = request.form.get("username","").strip()
        modules = request.form.getlist("modules")
        set_modules(username, modules)
        return render_template("users.html", users=list_users(), modules=MODULES, ok="Módulos actualizados.", error=None)
    except Exception as e:
        return render_template("users.html", users=list_users(), modules=MODULES, ok=None, error=str(e))

@app.route("/admin/users/password", methods=["POST"])
@admin_required
def admin_users_password():
    try:
        set_password(
            request.form.get("username","").strip(),
            request.form.get("password","").strip()
        )
        return render_template("users.html", users=list_users(), modules=MODULES, ok="Contraseña actualizada.", error=None)
    except Exception as e:
        return render_template("users.html", users=list_users(), modules=MODULES, ok=None, error=str(e))

@app.route("/admin/users/delete", methods=["POST"])
@admin_required
def admin_users_delete():
    try:
        delete_user(request.form.get("username","").strip())
        return render_template("users.html", users=list_users(), modules=MODULES, ok="Usuario eliminado.", error=None)
    except Exception as e:
        return render_template("users.html", users=list_users(), modules=MODULES, ok=None, error=str(e))

# =========================
# API FACTURACIÓN (PROTEGIDA) (existente)
# =========================
@app.route("/proveedores", methods=["GET"])
@login_required
@module_required("facturacion_patio")
def proveedores_get():
    prov_map = load_proveedores()
    out = []
    for nombre, info in sorted(prov_map.items(), key=lambda x: x[0]):
        item = {"nombre": nombre, "tipo": info.get("tipo", "hora")}
        if item["tipo"] == "dia":
            item["pago_dia"] = float(info.get("pago_dia", 0) or 0)
            item["pago_dia_f"] = float(info.get("pago_dia_f", 0) or 0)
        else:
            item["pago_h"] = float(info.get("pago_h", 0) or 0)
            item["pago_f"] = float(info.get("pago_f", 0) or 0)
        out.append(item)
    return jsonify({"proveedores": out})

@app.route("/proveedores", methods=["POST"])
@login_required
@module_required("facturacion_patio")
def proveedores_post():
    payload = request.json or {}
    nombre = str(payload.get("nombre", "")).strip()
    if not nombre:
        return jsonify({"error": "Falta nombre proveedor"}), 400
    data = payload.get("data", {}) or {}
    upsert_proveedor(nombre, data)
    return jsonify({"ok": True})

@app.route("/upload", methods=["POST"])
@login_required
@module_required("facturacion_patio")
def upload():
    operativa = request.form.get("operativa") or DEFAULT_OPERATIVA
    if operativa not in OPERATIVAS:
        operativa = DEFAULT_OPERATIVA
    cfg = OPERATIVAS[operativa]

    f = request.files.get("pdf")
    if not f:
        return jsonify({"error": "No se recibió PDF"}), 400

    filename = secure_filename(f.filename or "archivo.pdf")
    if not filename.lower().endswith(".pdf"):
        filename += ".pdf"

    path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    f.save(path)

    rows = parse_and_group(path)

    for r in rows:
        r["ConductorKey"] = key_name(r.get("Conductor", ""))
        r.setdefault("Ruta", cfg["default_ruta"])
        r.setdefault("Matricula", "")
        r.setdefault("OverrideCoste", False)
        r.setdefault("CosteManual", None)
        r.setdefault("MatriculaFromMaster", False)
        r.setdefault("Proveedor", "")

    master_map = load_master_map(cfg["master_xlsx"])
    rows = apply_master(rows, master_map, cfg["default_ruta"])

    ensure_vehiculos_master(cfg["vehiculos_xlsx"])
    veh_map = load_vehiculo_map(cfg["vehiculos_xlsx"])
    for r in rows:
        mat = key_plate(r.get("Matricula", ""))
        if mat and mat in veh_map:
            r["Proveedor"] = veh_map[mat]

    prov_map = load_proveedores()
    prov_list = sorted(list(prov_map.keys()))

    return jsonify({
        "rows": rows,
        "cobro_ruta": cfg["cobro_ruta"],
        "rutas": cfg["rutas"],
        "operativa": operativa,
        "proveedores": prov_list,
        "proveedores_full": prov_map,
    })

@app.route("/export", methods=["POST"])
@login_required
@module_required("facturacion_patio")
def export():
    payload = request.json or {}
    operativa = payload.get("operativa") or DEFAULT_OPERATIVA
    if operativa not in OPERATIVAS:
        operativa = DEFAULT_OPERATIVA
    cfg = OPERATIVAS[operativa]

    rows = payload.get("rows", [])
    date_iso = payload.get("date") or datetime.date.today().isoformat()
    es_festivo = bool(payload.get("festivo", False))

    template_xlsx = cfg["template_xlsx"]
    if not os.path.exists(template_xlsx):
        return jsonify({"error": f"No encuentro la plantilla: {template_xlsx}"}), 400

    prov_map = load_proveedores()

    try:
        out, total_horas, total_coste = generate_meribia_xlsx(
            rows, date_iso, template_xlsx, es_festivo, prov_map,
            cliente_codigo=cfg["cliente_codigo"],
            remolque_ref=cfg["remolque_ref"],
            proyecto=cfg["proyecto"],
            cobro_ruta=cfg["cobro_ruta"],
        )

        save_master_from_rows(cfg["master_xlsx"], rows)
        ensure_vehiculos_master(cfg["vehiculos_xlsx"])
        save_vehiculo_map(cfg["vehiculos_xlsx"], rows)

        append_kpi(date_iso, operativa, rows, cfg["cobro_ruta"], cfg["rutas"], es_festivo, total_horas, total_coste)

    except Exception as e:
        return jsonify({"error": str(e)}), 400

    return send_file(out, as_attachment=True)

@app.route("/kpis/json", methods=["GET"])
@login_required
@module_required("facturacion_patio")
def kpis_json():
    if not os.path.exists(KPI_FILE):
        return jsonify({"kpis": []})
    df = pd.read_csv(KPI_FILE, sep=";")
    return jsonify({"kpis": df.to_dict(orient="records")})

@app.route("/kpis/xlsx", methods=["GET"])
@login_required
@module_required("facturacion_patio")
def kpis_xlsx():
    if not os.path.exists(KPI_FILE):
        return jsonify({"error": "No hay KPIs todavía."}), 400
    df = pd.read_csv(KPI_FILE, sep=";")
    out = p("kpis_facturacion.xlsx")
    df.to_excel(out, index=False)
    return send_file(out, as_attachment=True)

@app.route("/me")
@login_required
def me():
    return jsonify({
        "username": session.get("username"),
        "role": session.get("role"),
        "modules": session.get("modules_list", []),
    })

# ============================================================
# ===================== NUEVO: DATA DB =======================
# ============================================================
DB_DATA = p("data.db")

def db_data():
    conn = sqlite3.connect(DB_DATA)
    conn.row_factory = sqlite3.Row
    return conn

def init_data_db():
    with db_data() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS clientes(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                codigo TEXT NOT NULL,
                nombre TEXT NOT NULL,
                cif TEXT DEFAULT '',
                direccion TEXT DEFAULT '',
                notas TEXT DEFAULT '',
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS clientes_contactos(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cliente_id INTEGER NOT NULL,
                nombre TEXT DEFAULT '',
                telefono TEXT DEFAULT '',
                email TEXT DEFAULT '',
                cargo TEXT DEFAULT '',
                FOREIGN KEY(cliente_id) REFERENCES clientes(id)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS proveedores_fichas(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT NOT NULL,
                cif TEXT DEFAULT '',
                direccion TEXT DEFAULT '',
                telefono TEXT DEFAULT '',
                email TEXT DEFAULT '',
                notas TEXT DEFAULT '',
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS vehiculos(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                tipo TEXT NOT NULL, -- TRACTORA/RIGIDO/SEMI_FRIO/SEMI_CAJA/SEMI_LONA/SEMI_PLATAFORMA/FURGONETA/OTRO
                matricula TEXT NOT NULL,
                proveedor TEXT DEFAULT '',
                itv_fecha TEXT DEFAULT '',
                tacografo_fecha TEXT DEFAULT '',
                atp_fecha TEXT DEFAULT '',
                notas TEXT DEFAULT '',
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS rutas(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                cliente_codigo TEXT DEFAULT 'ECI',
                cliente_nombre TEXT DEFAULT 'EL CORTE INGLES',
                proyecto TEXT DEFAULT '',
                fecha TEXT DEFAULT '',
                origen TEXT DEFAULT '',
                destino TEXT DEFAULT '',
                estado TEXT DEFAULT 'BORRADOR', -- BORRADOR / FINAL
                pdf_filename TEXT DEFAULT '',
                created_by TEXT DEFAULT '',
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS ruta_items(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                ruta_id INTEGER NOT NULL,
                tipo TEXT DEFAULT 'PROPIOS', -- AJENOS/PROPIOS/RETORNO/DOMINGO
                colaborador TEXT DEFAULT '',
                conductor TEXT DEFAULT '',
                vehiculo TEXT DEFAULT '',
                remolque TEXT DEFAULT '',
                origen TEXT DEFAULT '',
                destino TEXT DEFAULT '',
                salida TEXT DEFAULT '',
                llegada TEXT DEFAULT '',
                km_aprox REAL DEFAULT 0,
                duracion_h REAL DEFAULT 0,
                notas TEXT DEFAULT '',
                FOREIGN KEY(ruta_id) REFERENCES rutas(id)
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS planificacion(
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                titulo TEXT NOT NULL,
                payload_json TEXT NOT NULL,
                created_by TEXT DEFAULT '',
                created_at TEXT NOT NULL
            )
        """)
        conn.commit()

init_data_db()

# ============================================================
# ================= NUEVO: HELPERS RUTAS ======================
# ============================================================
TIPOS_VALIDOS = ["AJENOS", "PROPIOS", "RETORNO", "DOMINGO"]

def etiqueta_tipo(letter: str) -> str:
    m = (letter or "").strip().upper()
    if m == "A": return "AJENOS"
    if m == "X": return "PROPIOS"
    if m == "R": return "RETORNO"
    if m == "D": return "DOMINGO"
    return "PROPIOS"

def approx_km_and_time(origen: str, destino: str):
    # Placeholder (sin API externa): dejamos 0 y lo completaremos con integración (Google/OSRM) luego
    return 0.0, 0.0

def enforce_uniques(items: list):
    """
    Conductor+vehiculo NO repetidos salvo tipo=RETORNO.
    """
    used_conductores = set()
    used_vehiculos = set()
    errors = []

    for idx, it in enumerate(items):
        t = (it.get("tipo","") or "").upper().strip()
        c = (it.get("conductor","") or "").strip()
        v = (it.get("vehiculo","") or "").strip()

        if t == "RETORNO":
            continue

        if c and c in used_conductores:
            errors.append(f"Fila {idx+1}: Conductor repetido en IDA ({c})")
        if v and v in used_vehiculos:
            errors.append(f"Fila {idx+1}: Vehículo repetido en IDA ({v})")

        if c: used_conductores.add(c)
        if v: used_vehiculos.add(v)

    return errors

def parse_eci_route_pdf_basic(pdf_path: str):
    """
    Parser básico: extrae texto completo y devuelve líneas.
    Luego tú asignas/ajustas en la tabla.
    (Lo refinamos en siguientes iteraciones con tus PDFs reales.)
    """
    all_lines = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            txt = page.extract_text() or ""
            for ln in (txt.splitlines() if txt else []):
                ln = ln.strip()
                if ln:
                    all_lines.append(ln)

    # intentamos detectar fecha en algo tipo 08/02/2026
    fecha = ""
    for ln in all_lines[:40]:
        m = re.search(r"(\d{2}/\d{2}/\d{4})", ln)
        if m:
            try:
                d = datetime.datetime.strptime(m.group(1), "%d/%m/%Y").date()
                fecha = d.isoformat()
                break
            except Exception:
                pass

    return {"fecha": fecha, "raw_lines": all_lines[:400]}

# ============================================================
# ====================== NUEVO: RUTAS UI ======================
# ============================================================
@app.route("/rutas")
@login_required
@module_required("rutas")
def rutas_home():
    with db_data() as conn:
        rutas = conn.execute("SELECT * FROM rutas ORDER BY id DESC LIMIT 200").fetchall()
    return render_template("rutas.html", rutas=rutas)

@app.route("/rutas/new", methods=["POST"])
@login_required
@module_required("rutas")
def rutas_new():
    titulo = (request.form.get("titulo") or "").strip() or "Ruta sin título"
    proyecto = (request.form.get("proyecto") or "").strip()
    with db_data() as conn:
        cur = conn.execute(
            "INSERT INTO rutas(cliente_codigo,cliente_nombre,proyecto,estado,created_by,created_at) VALUES(?,?,?,?,?,?)",
            ("ECI","EL CORTE INGLES", proyecto, "BORRADOR", session.get("username",""), datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        ruta_id = cur.lastrowid
        conn.commit()
    return redirect(url_for("rutas_edit", ruta_id=ruta_id))

@app.route("/rutas/<int:ruta_id>")
@login_required
@module_required("rutas")
def rutas_edit(ruta_id: int):
    with db_data() as conn:
        ruta = conn.execute("SELECT * FROM rutas WHERE id=?", (ruta_id,)).fetchone()
        items = conn.execute("SELECT * FROM ruta_items WHERE ruta_id=? ORDER BY id ASC", (ruta_id,)).fetchall()
    if not ruta:
        return "Ruta no existe", 404
    return render_template("ruta_edit.html", ruta=ruta, items=items, tipos=TIPOS_VALIDOS)

@app.route("/rutas/<int:ruta_id>/upload_pdf", methods=["POST"])
@login_required
@module_required("rutas")
@app.route("/rutas/upload_pdf", methods=["POST"])
@login_required
@module_required("rutas")
def rutas_upload_pdf_alias():
    """
    Alias para soportar POST /rutas/upload_pdf.
    El frontend debe mandar route_id (por ejemplo 3) en el form.
    """
    rid = (request.form.get("route_id") or "").strip()
    if not rid.isdigit():
        return jsonify({"error": "Falta route_id o no es válido"}), 400

    return rutas_upload_pdf(int(rid))
def rutas_upload_pdf(ruta_id: int):
    f = request.files.get("pdf")
    if not f:
        return redirect(url_for("rutas_edit", ruta_id=ruta_id))

    filename = secure_filename(f.filename or "ruta.pdf")
    if not filename.lower().endswith(".pdf"):
        filename += ".pdf"
    path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    f.save(path)

    parsed = parse_eci_route_pdf_basic(path)
    fecha = parsed.get("fecha","")

    # Creamos items "vacíos" a partir de algunas líneas detectadas (MVP)
    raw = parsed.get("raw_lines", [])
    suggested = []
    for ln in raw:
        # Si ves algún patrón claro luego, lo refinamos.
        # De momento: no inventamos rutas.
        if "ORIGEN" in ln.upper() or "DESTINO" in ln.upper():
            continue

    with db_data() as conn:
        conn.execute("UPDATE rutas SET pdf_filename=?, fecha=? WHERE id=?", (filename, fecha, ruta_id))
        conn.commit()

    return redirect(url_for("rutas_edit", ruta_id=ruta_id))

@app.route("/rutas/<int:ruta_id>/save", methods=["POST"])
@login_required
@module_required("rutas")
def rutas_save(ruta_id: int):
    payload = request.json or {}
    ruta = payload.get("ruta") or {}
    items = payload.get("items") or []

    # normaliza tipos
    for it in items:
        t = (it.get("tipo") or "PROPIOS").upper().strip()
        if t not in TIPOS_VALIDOS:
            t = "PROPIOS"
        it["tipo"] = t

    errors = enforce_uniques(items)
    if errors:
        return jsonify({"error": "\n".join(errors)}), 400

    with db_data() as conn:
        conn.execute(
            "UPDATE rutas SET proyecto=?, fecha=?, origen=?, destino=? WHERE id=?",
            (
                (ruta.get("proyecto") or "").strip(),
                (ruta.get("fecha") or "").strip(),
                (ruta.get("origen") or "").strip(),
                (ruta.get("destino") or "").strip(),
                ruta_id
            )
        )
        conn.execute("DELETE FROM ruta_items WHERE ruta_id=?", (ruta_id,))
        for it in items:
            conn.execute("""
                INSERT INTO ruta_items(
                    ruta_id,tipo,colaborador,conductor,vehiculo,remolque,origen,destino,salida,llegada,km_aprox,duracion_h,notas
                ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
            """, (
                ruta_id,
                it.get("tipo","PROPIOS"),
                (it.get("colaborador") or "").strip(),
                (it.get("conductor") or "").strip(),
                (it.get("vehiculo") or "").strip(),
                (it.get("remolque") or "").strip(),
                (it.get("origen") or "").strip(),
                (it.get("destino") or "").strip(),
                (it.get("salida") or "").strip(),
                (it.get("llegada") or "").strip(),
                float(it.get("km_aprox") or 0),
                float(it.get("duracion_h") or 0),
                (it.get("notas") or "").strip(),
            ))
        conn.commit()

    return jsonify({"ok": True})

@app.route("/rutas/<int:ruta_id>/to_planificacion", methods=["POST"])
@login_required
@module_required("rutas")
def rutas_to_planificacion(ruta_id: int):
    with db_data() as conn:
        ruta = conn.execute("SELECT * FROM rutas WHERE id=?", (ruta_id,)).fetchone()
        items = conn.execute("SELECT * FROM ruta_items WHERE ruta_id=? ORDER BY id ASC", (ruta_id,)).fetchall()
    if not ruta:
        return jsonify({"error":"Ruta no existe"}), 404

    pack = {
        "ruta": dict(ruta),
        "items": [dict(x) for x in items],
        "saved_at": datetime.datetime.now().isoformat(timespec="seconds")
    }

    titulo = f"PLANIF · Ruta #{ruta_id} · {ruta['fecha'] or 'sin_fecha'}"
    with db_data() as conn:
        conn.execute(
            "INSERT INTO planificacion(titulo,payload_json,created_by,created_at) VALUES(?,?,?,?)",
            (titulo, json.dumps(pack, ensure_ascii=False), session.get("username",""), datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()

    return jsonify({"ok": True})

# ============================================================
# ==================== NUEVO: PLANIFICACIÓN ==================
# ============================================================
@app.route("/planificacion")
@login_required
@module_required("planificacion")
def planificacion_home():
    with db_data() as conn:
        rows = conn.execute("SELECT id,titulo,created_by,created_at FROM planificacion ORDER BY id DESC LIMIT 200").fetchall()
    return render_template("planificacion.html", rows=rows)

@app.route("/planificacion/<int:pid>")
@login_required
@module_required("planificacion")
def planificacion_view(pid: int):
    with db_data() as conn:
        row = conn.execute("SELECT * FROM planificacion WHERE id=?", (pid,)).fetchone()
    if not row:
        return "No existe", 404
    payload = json.loads(row["payload_json"])
    return render_template("ruta_edit.html", ruta=payload["ruta"], items=payload["items"], tipos=TIPOS_VALIDOS, readonly=True)

# ============================================================
# ===================== NUEVO: CLIENTES ======================
# ============================================================
@app.route("/clientes")
@login_required
@module_required("clientes")
def clientes_home():
    with db_data() as conn:
        clientes = conn.execute("SELECT * FROM clientes ORDER BY id DESC LIMIT 300").fetchall()
    return render_template("clientes.html", clientes=clientes)

@app.route("/clientes/create", methods=["POST"])
@login_required
@module_required("clientes")
def clientes_create():
    codigo = (request.form.get("codigo") or "").strip()
    nombre = (request.form.get("nombre") or "").strip()
    cif = (request.form.get("cif") or "").strip()
    direccion = (request.form.get("direccion") or "").strip()
    if not codigo or not nombre:
        return redirect(url_for("clientes_home"))
    with db_data() as conn:
        conn.execute(
            "INSERT INTO clientes(codigo,nombre,cif,direccion,notas,created_at) VALUES(?,?,?,?,?,?)",
            (codigo, nombre, cif, direccion, "", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()
    return redirect(url_for("clientes_home"))

# ============================================================
# ==================== NUEVO: PROVEEDORES ====================
# ============================================================
@app.route("/proveedores_fichas")
@login_required
@module_required("proveedores_fichas")
def proveedores_fichas_home():
    with db_data() as conn:
        provs = conn.execute("SELECT * FROM proveedores_fichas ORDER BY id DESC LIMIT 300").fetchall()
    return render_template("proveedores_ficha.html", provs=provs)

@app.route("/proveedores_fichas/create", methods=["POST"])
@login_required
@module_required("proveedores_fichas")
def proveedores_fichas_create():
    nombre = (request.form.get("nombre") or "").strip()
    if not nombre:
        return redirect(url_for("proveedores_fichas_home"))
    cif = (request.form.get("cif") or "").strip()
    direccion = (request.form.get("direccion") or "").strip()
    telefono = (request.form.get("telefono") or "").strip()
    email = (request.form.get("email") or "").strip()
    with db_data() as conn:
        conn.execute(
            "INSERT INTO proveedores_fichas(nombre,cif,direccion,telefono,email,notas,created_at) VALUES(?,?,?,?,?,?,?)",
            (nombre, cif, direccion, telefono, email, "", datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        )
        conn.commit()
    return redirect(url_for("proveedores_fichas_home"))

# ============================================================
# ====================== NUEVO: VEHÍCULOS =====================
# ============================================================
@app.route("/vehiculos")
@login_required
@module_required("vehiculos")
def vehiculos_home():
    with db_data() as conn:
        vehs = conn.execute("SELECT * FROM vehiculos ORDER BY id DESC LIMIT 500").fetchall()
    return render_template("vehiculos.html", vehs=vehs)

@app.route("/vehiculos/create", methods=["POST"])
@login_required
@module_required("vehiculos")
def vehiculos_create():
    tipo = (request.form.get("tipo") or "").strip().upper()
    matricula = (request.form.get("matricula") or "").strip().upper()
    if not tipo or not matricula:
        return redirect(url_for("vehiculos_home"))
    proveedor = (request.form.get("proveedor") or "").strip()
    itv = (request.form.get("itv_fecha") or "").strip()
    tac = (request.form.get("tacografo_fecha") or "").strip()
    atp = (request.form.get("atp_fecha") or "").strip()
    notas = (request.form.get("notas") or "").strip()
    with db_data() as conn:
        conn.execute("""
            INSERT INTO vehiculos(tipo,matricula,proveedor,itv_fecha,tacografo_fecha,atp_fecha,notas,created_at)
            VALUES(?,?,?,?,?,?,?,?)
        """, (
            tipo, matricula, proveedor, itv, tac, atp, notas,
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ))
        conn.commit()
    return redirect(url_for("vehiculos_home"))

# ============================================================
# START
# ============================================================
if __name__ == "__main__":
    init_users_db()
    ensure_default_admin()
    init_data_db()
    app.run(host="0.0.0.0", port=5000, debug=True)